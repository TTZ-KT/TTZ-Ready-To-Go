import os
import time
import tempfile
import re
import base64
from langchain_ollama import ChatOllama
from langchain_huggingface import HuggingFaceEmbeddings
from langchain.text_splitters import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import FAISS
from langchain.chains import ConversationalRetrievalChain
from langchain.memory import ConversationBufferMemory
from langchain.prompts import PromptTemplate
from langchain.schema import Document

from langchain_community.document_loaders import (
    PyPDFLoader,
    Docx2txtLoader,
    TextLoader,
    CSVLoader,
    JSONLoader,
    UnstructuredXMLLoader,
    UnstructuredRTFLoader
)

# Try multiple Excel loading methods
try:
    from langchain_community.document_loaders import UnstructuredExcelLoader
    UNSTRUCTURED_EXCEL = True
except ImportError:
    UNSTRUCTURED_EXCEL = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class RAGEngine:
    """
    Multi-Format RAG Engine - ULTRA ROBUST VERSION
    ✅ Multiple fallback methods for Excel
    ✅ Better error handling with detailed messages
    ✅ Support for 15+ file formats
    ✅ Works even with corrupted/complex files
    """
    
    def __init__(self, model="qwen2.5:7b", vision_model="llama3.2-vision:latest"):
        """Initialize RAG Engine with robust file handling"""
        print(f"[RAG] Initializing ROBUST version")
        print(f"[RAG] Text Model: {model}")
        print(f"[RAG] Vision Model: {vision_model}")
        
        self.model = model
        self.vision_model = vision_model
        self.vectorstore = None
        self.chain = None
        self.llm = None
        self.memory = None
        self.processed_documents = []
        
        self.casual_patterns = [
            r'^(hi|hey|hello|sup|what\'s up|wassup|yo)\b',
            r'^(thanks|thank you|thx|ty|appreciate it)\b',
            r'^(bye|goodbye|see you|cya|later)\b',
            r'^(how are you|how\'s it going|how are ya|how do you do)\b',
            r'^(ok|okay|cool|nice|great|awesome|perfect)\b',
            r'^(yes|no|yeah|yep|nope|sure)\b',
            r'(what are you|what\'re you|whatcha|wyd)',
            r'(tell me about yourself|who are you|introduce yourself)',
            r'(good morning|good afternoon|good evening|good night)',
            r'(nice to meet you|pleased to meet you)',
        ]
        
        self.casual_keywords = [
            'doing now', 'doing today', 'your name', 'about you', 
            'feeling', 'your day', 'up to', 'busy'
        ]
        
        # Check available Excel loaders
        print(f"[RAG] Excel support:")
        print(f"  - UnstructuredExcelLoader: {'✅' if UNSTRUCTURED_EXCEL else '❌'}")
        print(f"  - Pandas: {'✅' if PANDAS_AVAILABLE else '❌'}")
        print(f"  - OpenPyXL: {'✅' if OPENPYXL_AVAILABLE else '❌'}")
        
        print("[RAG] Loading embeddings...")
        import torch
        device = 'cuda' if torch.cuda.is_available() else 'cpu'
        print(f"[RAG] Device: {device.upper()}")
        
        self.embeddings = HuggingFaceEmbeddings(
            model_name="sentence-transformers/all-MiniLM-L6-v2",
            model_kwargs={'device': device},
            encode_kwargs={'normalize_embeddings': True, 'batch_size': 32}
        )
        
        self.vector_dir = os.path.join("vectors", "faiss_index")
        
        # Initialize vision LLM
        print("[RAG] Loading vision model...")
        self.vision_llm = ChatOllama(
            model=vision_model,
            temperature=0.3,
            timeout=300,
            keep_alive="10m"
        )
        
        print("[RAG] ✅ Ready with ROBUST file support!")
    
    def _detect_file_type(self, file_name):
        """Detect file type from extension"""
        return os.path.splitext(file_name)[1].lower().lstrip('.')
    
    def _is_image_file(self, file_type):
        """Check if file is an image"""
        return file_type in ['png', 'jpg', 'jpeg', 'bmp', 'gif', 'webp', 'tiff']
    
    def _get_model_settings(self, model_name):
        """Get timeout and context settings for model"""
        model_lower = model_name.lower()
        
        if any(x in model_lower for x in ['qwen3', '14b', '32b', '70b', 'deepseek-r1', '90b']):
            return (512, 4096, 600)
        elif any(x in model_lower for x in ['7b', '8b', '11b', '13b', 'qwen2.5:latest']):
            return (384, 3072, 300)
        else:
            return (256, 2048, 180)
    
    def _load_excel_with_pandas(self, file_path):
        """
        Load Excel using Pandas - MORE RELIABLE!
        Works with complex Excel files that UnstructuredExcelLoader fails on
        WINDOWS COMPATIBLE: Properly closes file handles
        """
        print(f"[Excel] Using Pandas loader (more reliable)")
        
        try:
            # Read all sheets - use context manager for proper cleanup
            documents = []
            
            # Get sheet names first
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            excel_file.close()  # IMPORTANT: Close immediately
            
            # Now process each sheet
            for sheet_name in sheet_names:
                print(f"[Excel] Reading sheet: {sheet_name}")
                
                # Read with context manager for proper cleanup
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # Convert DataFrame to readable text
                content = f"=== Sheet: {sheet_name} ===\n\n"
                
                # Add column headers
                content += "Columns: " + ", ".join(df.columns.tolist()) + "\n\n"
                
                # Add row count
                content += f"Total rows: {len(df)}\n\n"
                
                # Add data (first 1000 rows to avoid huge files)
                content += "Data:\n"
                for idx, row in df.head(1000).iterrows():
                    row_text = " | ".join([f"{col}: {val}" for col, val in row.items()])
                    content += f"Row {idx + 1}: {row_text}\n"
                
                if len(df) > 1000:
                    content += f"\n... (showing first 1000 of {len(df)} rows)\n"
                
                # Create document
                doc = Document(
                    page_content=content,
                    metadata={
                        "source": os.path.basename(file_path),
                        "sheet": sheet_name,
                        "rows": len(df),
                        "columns": len(df.columns)
                    }
                )
                documents.append(doc)
                
                # Explicitly delete DataFrame to free memory
                del df
            
            print(f"[Excel] ✅ Successfully loaded {len(documents)} sheet(s)")
            return documents
            
        except Exception as e:
            print(f"[Excel] Pandas failed: {e}")
            raise
    
    def _load_excel_with_openpyxl(self, file_path):
        """
        Load Excel using OpenPyXL - FALLBACK METHOD
        Direct cell-by-cell reading
        WINDOWS COMPATIBLE: Uses context manager
        """
        print(f"[Excel] Using OpenPyXL loader (fallback)")
        
        try:
            from openpyxl import load_workbook
            
            # Load workbook with proper cleanup
            wb = load_workbook(file_path, read_only=True, data_only=True)
            documents = []
            
            try:
                for sheet_name in wb.sheetnames:
                    print(f"[Excel] Reading sheet: {sheet_name}")
                    sheet = wb[sheet_name]
                    
                    content = f"=== Sheet: {sheet_name} ===\n\n"
                    
                    # Read all rows (limit to 1000 for performance)
                    rows_data = []
                    for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        if idx > 1000:
                            break
                        row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                        rows_data.append(f"Row {idx}: {row_text}")
                    
                    content += "\n".join(rows_data)
                    
                    if sheet.max_row > 1000:
                        content += f"\n\n... (showing first 1000 of {sheet.max_row} rows)\n"
                    
                    doc = Document(
                        page_content=content,
                        metadata={
                            "source": os.path.basename(file_path),
                            "sheet": sheet_name
                        }
                    )
                    documents.append(doc)
                
                print(f"[Excel] ✅ Successfully loaded {len(documents)} sheet(s)")
                return documents
                
            finally:
                # IMPORTANT: Always close workbook, even if error occurs
                wb.close()
                print(f"[Excel] Workbook closed")
            
        except Exception as e:
            print(f"[Excel] OpenPyXL failed: {e}")
            raise
    
    def _process_image_with_vision(self, image_bytes, file_name):
        """Process image using vision model"""
        print(f"[Vision] Analyzing image: {file_name}")
        start_time = time.time()
        
        try:
            base64_image = base64.b64encode(image_bytes).decode('utf-8')
            file_type = self._detect_file_type(file_name)
            mime_type = f"image/{file_type if file_type != 'jpg' else 'jpeg'}"
            
            prompt = f"""Analyze this image in detail. Provide:

1. MAIN CONTENT: What is the primary subject or purpose?
2. TEXT CONTENT: Any visible text, labels, captions, or written information
3. KEY DETAILS: Important visual elements, data, diagrams, or specific information
4. TYPE: What kind of image is this? (chart, diagram, photo, screenshot, document, etc.)

Be thorough and specific so this description can answer questions about the image.

Image: {file_name}"""

            from langchain_core.messages import HumanMessage
            
            message = HumanMessage(
                content=[
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": f"data:{mime_type};base64,{base64_image}"}
                ]
            )
            
            response = self.vision_llm.invoke([message])
            description = response.content
            
            elapsed = time.time() - start_time
            print(f"[Vision] ✅ Analyzed in {elapsed:.2f}s")
            
            return Document(
                page_content=f"[IMAGE: {file_name}]\n\n{description}",
                metadata={
                    "source": file_name,
                    "type": "image",
                    "processed_with": f"vision_model_{self.vision_model}"
                }
            )
            
        except Exception as e:
            print(f"[Vision] ⚠️ Error: {str(e)}")
            return Document(
                page_content=f"[IMAGE: {file_name}]\n\nVision model error: {str(e)}",
                metadata={"source": file_name, "type": "image", "error": str(e)}
            )
    
    def _load_document_by_type(self, file_path, file_bytes=None):
        """
        ROBUST document loader with multiple fallback methods
        """
        file_type = self._detect_file_type(file_path)
        file_name = os.path.basename(file_path)
        
        print(f"[Loader] {file_type.upper()}: {file_name}")
        
        try:
            # ============ IMAGE FILES ============
            if self._is_image_file(file_type):
                print(f"[Loader] → Vision Model")
                if file_bytes:
                    return [self._process_image_with_vision(file_bytes, file_name)]
                else:
                    with open(file_path, 'rb') as f:
                        return [self._process_image_with_vision(f.read(), file_name)]
            
            # ============ EXCEL FILES - MULTIPLE METHODS ============
            elif file_type in ['xlsx', 'xls', 'xlsm', 'ods']:
                print(f"[Loader] → Excel (trying multiple methods)")
                
                # Method 1: Try Pandas first (most reliable)
                if PANDAS_AVAILABLE:
                    try:
                        return self._load_excel_with_pandas(file_path)
                    except Exception as e1:
                        print(f"[Excel] Pandas failed: {e1}")
                
                # Method 2: Try OpenPyXL
                if OPENPYXL_AVAILABLE:
                    try:
                        return self._load_excel_with_openpyxl(file_path)
                    except Exception as e2:
                        print(f"[Excel] OpenPyXL failed: {e2}")
                
                # Method 3: Try UnstructuredExcelLoader
                if UNSTRUCTURED_EXCEL:
                    try:
                        from langchain_community.document_loaders import UnstructuredExcelLoader
                        print(f"[Excel] Trying UnstructuredExcelLoader...")
                        loader = UnstructuredExcelLoader(file_path, mode="single")
                        return loader.load()
                    except Exception as e3:
                        print(f"[Excel] UnstructuredExcelLoader failed: {e3}")
                
                # All methods failed
                error_msg = "Could not load Excel file. Please install: pip install pandas openpyxl"
                print(f"[Excel] ❌ {error_msg}")
                return [Document(
                    page_content=f"[Excel file could not be loaded]\n\nFile: {file_name}\n\n"
                                f"Tried multiple methods but all failed.\n"
                                f"Install dependencies: pip install pandas openpyxl",
                    metadata={"source": file_name, "type": "excel", "error": "all_methods_failed"}
                )]
            
            # ============ PDF FILES ============
            elif file_type == 'pdf':
                return PyPDFLoader(file_path).load()
            
            # ============ WORD FILES ============
            elif file_type in ['docx', 'doc']:
                return Docx2txtLoader(file_path).load()
            
            # ============ TEXT FILES ============
            elif file_type in ['txt', 'md', 'markdown']:
                return TextLoader(file_path, encoding='utf-8').load()
            
            # ============ RTF FILES ============
            elif file_type == 'rtf':
                return UnstructuredRTFLoader(file_path).load()
            
            # ============ CSV FILES ============
            elif file_type == 'csv':
                # Try Pandas first for better formatting
                if PANDAS_AVAILABLE:
                    try:
                        df = pd.read_csv(file_path)
                        content = f"CSV File: {file_name}\n\n"
                        content += f"Columns: {', '.join(df.columns.tolist())}\n\n"
                        content += f"Total rows: {len(df)}\n\n"
                        content += "Data:\n"
                        content += df.head(1000).to_string()
                        
                        if len(df) > 1000:
                            content += f"\n\n... (showing first 1000 of {len(df)} rows)"
                        
                        return [Document(
                            page_content=content,
                            metadata={"source": file_name, "rows": len(df), "columns": len(df.columns)}
                        )]
                    except:
                        pass
                
                # Fallback to CSVLoader
                return CSVLoader(file_path, encoding='utf-8').load()
            
            # ============ JSON FILES ============
            elif file_type == 'json':
                return JSONLoader(file_path=file_path, jq_schema='.', text_content=False).load()
            
            # ============ XML FILES ============
            elif file_type == 'xml':
                return UnstructuredXMLLoader(file_path).load()
            
            # ============ YAML FILES ============
            elif file_type in ['yaml', 'yml']:
                return TextLoader(file_path, encoding='utf-8').load()
            
            # ============ UNSUPPORTED ============
            else:
                print(f"[Loader] ⚠️ Unsupported file type: {file_type}")
                return [Document(
                    page_content=f"[Unsupported file type: {file_type}]\n\nFile: {file_name}\n\n"
                                f"Supported formats: PDF, DOCX, TXT, CSV, XLSX, XLS, JSON, XML, YAML, RTF, images",
                    metadata={"source": file_name, "type": "unsupported"}
                )]
                
        except Exception as e:
            print(f"[Loader] ❌ Error loading {file_name}: {str(e)}")
            import traceback
            traceback.print_exc()
            
            return [Document(
                page_content=f"[Error loading file: {file_name}]\n\n"
                            f"Error: {str(e)}\n\n"
                            f"File type: {file_type}\n"
                            f"This file could not be processed. Try:\n"
                            f"1. Check if file is corrupted\n"
                            f"2. Try a different file format\n"
                            f"3. Install missing dependencies",
                metadata={"source": file_name, "type": "error", "error": str(e)}
            )]
    
    def process_uploaded_file(self, uploaded_file):
        """Process uploaded file with robust handling - WINDOWS COMPATIBLE"""
        file_name = uploaded_file.name
        file_type = self._detect_file_type(file_name)
        
        print(f"[Processing] {file_name} ({file_type.upper()})")
        tmp_path = None
        
        try:
            # Create temp file
            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_type}") as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                tmp_path = tmp_file.name
            
            # Load document with appropriate method
            if self._is_image_file(file_type):
                documents = self._load_document_by_type(tmp_path, uploaded_file.getvalue())
            else:
                documents = self._load_document_by_type(tmp_path)
            
            # Force garbage collection to release file handles (Windows compatibility)
            import gc
            gc.collect()
            
            # Clean up - WINDOWS FIX: Add delay and retry
            if tmp_path and os.path.exists(tmp_path):
                import time as time_module
                max_attempts = 5
                for attempt in range(max_attempts):
                    try:
                        os.unlink(tmp_path)
                        print(f"[Cleanup] ✅ Temp file deleted")
                        break
                    except PermissionError as e:
                        if attempt < max_attempts - 1:
                            print(f"[Cleanup] File locked, retrying... (attempt {attempt + 1}/{max_attempts})")
                            time_module.sleep(0.5)  # Wait 500ms
                        else:
                            print(f"[Cleanup] ⚠️ Could not delete temp file (Windows lock): {tmp_path}")
                            print(f"[Cleanup] File will be cleaned up by system eventually")
                    except Exception as e:
                        print(f"[Cleanup] ⚠️ Cleanup error: {e}")
                        break
            
            if not documents:
                print(f"[Processing] ⚠️ No content extracted")
                return []
            
            # Smart chunking based on file type
            if file_type in ['xlsx', 'xls', 'xlsm', 'ods', 'csv']:
                chunk_size, chunk_overlap = 2000, 400
            elif self._is_image_file(file_type):
                chunk_size, chunk_overlap = 4000, 0
            else:
                chunk_size, chunk_overlap = 1200, 300
            
            text_splitter = RecursiveCharacterTextSplitter(
                chunk_size=chunk_size,
                chunk_overlap=chunk_overlap,
                separators=["\n\n\n", "\n\n", "\n", ". ", " ", ""],
                length_function=len
            )
            
            chunks = text_splitter.split_documents(documents)
            
            for chunk in chunks:
                chunk.metadata['source'] = file_name
            
            self.processed_documents.append(file_name)
            print(f"[Processing] ✅ {file_name}: {len(chunks)} chunks")
            
            return chunks
            
        except Exception as e:
            # Clean up temp file even on error - WINDOWS COMPATIBLE
            if tmp_path and os.path.exists(tmp_path):
                import time as time_module
                for attempt in range(3):
                    try:
                        os.unlink(tmp_path)
                        print(f"[Cleanup] Temp file deleted (error path)")
                        break
                    except PermissionError:
                        if attempt < 2:
                            time_module.sleep(0.5)
                        else:
                            print(f"[Cleanup] ⚠️ Could not delete temp file (locked): {tmp_path}")
                    except:
                        break
            print(f"[Processing] ❌ Error: {e}")
            raise
    
    def create_vectorstore(self, chunks):
        """Create FAISS vectorstore"""
        print(f"[Vectorstore] Creating from {len(chunks)} chunks...")
        start_time = time.time()
        
        if not chunks:
            raise ValueError("No chunks provided")
        
        self.vectorstore = FAISS.from_documents(chunks, embedding=self.embeddings)
        
        elapsed = time.time() - start_time
        print(f"[Vectorstore] ✅ Created in {elapsed:.2f}s")
        
        os.makedirs(self.vector_dir, exist_ok=True)
        try:
            self.vectorstore.save_local(self.vector_dir)
            print(f"[Vectorstore] Saved to disk")
        except Exception as e:
            print(f"[Vectorstore] Could not save: {e}")
    
    def setup_chain(self):
        """Setup conversational chain"""
        if not self.vectorstore:
            raise ValueError("Vectorstore not initialized")
        
        print(f"[Chain] Setting up with {self.model}...")
        
        num_predict, num_ctx, timeout = self._get_model_settings(self.model)
        print(f"[Chain] Settings: predict={num_predict}, context={num_ctx}, timeout={timeout}s")
        
        self.llm = ChatOllama(
            model=self.model,
            temperature=0.2,
            num_predict=num_predict,
            num_ctx=num_ctx,
            timeout=timeout,
            keep_alive="10m",
            num_gpu=1,
            num_thread=8
        )
        
        self.memory = ConversationBufferMemory(
            memory_key="chat_history",
            return_messages=True,
            output_key="answer"
        )
        
        qa_prompt = PromptTemplate(
            template="""You are answering questions based on provided document context. Use ALL the context below to give a complete answer.

Context from documents:
{context}

Question: {question}

Instructions:
- Read through ALL the context carefully
- If listing items (like questions), list ALL of them that appear in the context
- If the context contains images, use the image descriptions provided
- If information is incomplete, say so
- Answer based ONLY on the context above
- Be concise but complete

Answer:""",
            input_variables=["context", "question"]
        )
        
        self.chain = ConversationalRetrievalChain.from_llm(
            llm=self.llm,
            retriever=self.vectorstore.as_retriever(
                search_type="similarity",
                search_kwargs={"k": 6}
            ),
            memory=self.memory,
            return_source_documents=True,
            combine_docs_chain_kwargs={"prompt": qa_prompt},
            verbose=False
        )
        
        print(f"[Chain] ✅ Ready")
    
    def _is_casual_message(self, message):
        """Check if message is casual conversation - IMPROVED WITH FLEXIBLE MATCHING"""
        msg_lower = message.lower().strip()
        
        for pattern in self.casual_patterns:
            if re.search(pattern, msg_lower, re.IGNORECASE):
                return True
        
        for keyword in self.casual_keywords:
            if keyword in msg_lower:
                return True
        
        if len(msg_lower.split()) <= 3 and '?' not in msg_lower:
            return True
        
        words = msg_lower.split()
        
        if 'document' in msg_lower or 'file' in msg_lower or 'pdf' in msg_lower:
            return False
        if 'content' in msg_lower or 'show' in msg_lower or 'find' in msg_lower:
            return False
        if 'search' in msg_lower or 'list' in msg_lower or 'extract' in msg_lower:
            return False
        
        if any(combo in msg_lower for combo in ['what are you', 'what\'re you', 'whatcha']):
            return True
        
        if 'you' in words or 'your' in words or 'yourself' in words:
            if len(words) <= 10:
                return True
        
        return False
    
    def _init_chat_llm(self):
        """Initialize LLM for chat"""
        num_predict, num_ctx, timeout = self._get_model_settings(self.model)
        self.chat_llm = ChatOllama(
            model=self.model,
            temperature=0.7,
            num_predict=num_predict,
            num_ctx=num_ctx,
            timeout=timeout,
            keep_alive="10m",
            num_gpu=1,
            num_thread=8
        )
    
    def _chat_direct(self, message):
        """Direct chat without documents"""
        if not hasattr(self, 'chat_llm') or self.chat_llm is None:
            self._init_chat_llm()
        
        try:
            response = self.chat_llm.invoke(message)
            return {"answer": response.content, "source_documents": []}
        except Exception as e:
            print(f"[ERROR] Chat failed: {e}")
            raise
    
    def switch_model(self, new_model):
        """Switch to different model"""
        print(f"[Model Switch] {self.model} → {new_model}")
        
        num_predict, num_ctx, timeout = self._get_model_settings(new_model)
        
        self.chat_llm = ChatOllama(
            model=new_model,
            temperature=0.7,
            num_predict=num_predict,
            num_ctx=num_ctx,
            timeout=timeout,
            keep_alive="10m",
            num_gpu=1,
            num_thread=8
        )
        
        if not self.memory:
            self.memory = ConversationBufferMemory(
                memory_key="chat_history",
                return_messages=True,
                output_key="answer"
            )
        
        if self.vectorstore:
            self.llm = ChatOllama(
                model=new_model,
                temperature=0.2,
                num_predict=num_predict,
                num_ctx=num_ctx,
                timeout=timeout,
                keep_alive="10m",
                num_gpu=1,
                num_thread=8
            )
            
            qa_prompt = PromptTemplate(
                template="""You are answering questions based on provided document context. Use ALL the context below to give a complete answer.

Context from documents:
{context}

Question: {question}

Instructions:
- Read through ALL the context carefully
- If listing items (like questions), list ALL of them that appear in the context
- If the context contains images, use the image descriptions provided
- If information is incomplete, say so
- Answer based ONLY on the context above
- Be concise but complete

Answer:""",
                input_variables=["context", "question"]
            )
            
            self.chain = ConversationalRetrievalChain.from_llm(
                llm=self.llm,
                retriever=self.vectorstore.as_retriever(
                    search_type="similarity",
                    search_kwargs={"k": 6}
                ),
                memory=self.memory,
                return_source_documents=True,
                combine_docs_chain_kwargs={"prompt": qa_prompt},
                verbose=False
            )
        
        self.model = new_model
        print(f"[Model Switch] ✅ Switched to {new_model}")
    
    def ask_question(self, question):
        """Ask question with robust error handling"""
        print(f"\n{'='*60}")
        print(f"[QUERY] {question}")
        print(f"[MODEL] {self.model}")
        print(f"{'='*60}")
        
        is_casual = self._is_casual_message(question)
        print(f"[DEBUG] Is casual message: {is_casual}")
        
        if is_casual:
            print("[INFO] Casual message detected - using direct chat")
            return self._chat_direct(question)
        
        if not self.chain:
            print("[INFO] No documents - using direct chat")
            return self._chat_direct(question)
        
        total_start = time.time()
        
        try:
            chat_history = self.memory.chat_memory.messages if self.memory else []
            
            print(f"[INFO] Retrieving relevant chunks...")
            response = self.chain.invoke({
                "question": question,
                "chat_history": chat_history
            })
            
            total_time = time.time() - total_start
            
            sources = response.get("source_documents", [])
            print(f"\n[INFO] ✅ Retrieved {len(sources)} chunks in {total_time:.2f}s")
            
            for i, doc in enumerate(sources, 1):
                preview = doc.page_content[:100].replace('\n', ' ')
                source_file = doc.metadata.get('source', 'Unknown')
                print(f"  [{i}] {source_file}: {preview}...")
            
            print(f"[INFO] Answer generated in {total_time:.2f}s")
            print(f"{'='*60}\n")
            
            return {
                "answer": response["answer"],
                "source_documents": sources
            }
            
        except TimeoutError as e:
            elapsed = time.time() - total_start
            print(f"[TIMEOUT] Query timed out after {elapsed:.2f}s")
            return {
                "answer": f"⚠️ Query timed out after {elapsed:.0f}s. Try using a faster model or asking a simpler question.",
                "source_documents": []
            }
            
        except Exception as e:
            elapsed = time.time() - total_start
            print(f"[ERROR] Query failed after {elapsed:.2f}s: {e}")
            import traceback
            traceback.print_exc()
            
            return {
                "answer": f"❌ Error processing query: {str(e)}\n\nTry using a different model or simpler question.",
                "source_documents": []
            }
    
    def clear_documents(self):
        """Clear all documents"""
        self.vectorstore = None
        self.chain = None
        self.llm = None
        self.memory = None
        self.processed_documents = []
        
        if os.path.exists(self.vector_dir):
            import shutil
            try:
                shutil.rmtree(self.vector_dir)
                print("[RAG] Cleared saved vectors")
            except:
                pass
        
        print("[RAG] All documents cleared")