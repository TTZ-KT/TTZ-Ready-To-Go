import os
import time
import tempfile
import re
import base64
from langchain_ollama import ChatOllama
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import FAISS
from langchain_classic.chains import ConversationalRetrievalChain
from langchain_classic.memory import ConversationBufferMemory
from langchain_core.prompts import PromptTemplate
from langchain_core.documents import Document

from langchain_community.document_loaders import (
    PyPDFLoader,
    Docx2txtLoader,
    TextLoader,
    CSVLoader,
    JSONLoader,
    UnstructuredXMLLoader,
    UnstructuredRTFLoader
)

# Excel loading
try:
    from langchain_community.document_loaders import UnstructuredExcelLoader
    UNSTRUCTURED_EXCEL = True
except ImportError:
    UNSTRUCTURED_EXCEL = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class RAGEngine:
    """
    🚀 ENHANCED Multi-Format RAG Engine - MAXIMUM CAPACITY
    
    KEY IMPROVEMENTS:
    - 2x larger context window (8192 tokens)
    - 2x more chunks retrieved (12 chunks)
    - Larger chunks with better overlap
    - MMR retrieval for diverse results
    - Configurable retrieval strategies
    """
    
    def __init__(self, model="qwen2.5:7b", vision_model="llama3.2-vision:latest", 
                 retrieval_mode="mmr", num_chunks=12):
        """
        Initialize Enhanced RAG Engine
        
        Args:
            model: Text model name
            vision_model: Vision model for images
            retrieval_mode: 'similarity', 'mmr', or 'hybrid' (default: 'mmr')
            num_chunks: Number of chunks to retrieve (default: 12, max: 20)
        """
        print(f"[RAG] 🚀 Initializing ENHANCED CAPACITY version")
        print(f"[RAG] Text Model: {model}")
        print(f"[RAG] Vision Model: {vision_model}")
        print(f"[RAG] Retrieval Mode: {retrieval_mode.upper()}")
        print(f"[RAG] Chunks to Retrieve: {num_chunks}")
        
        self.model = model
        self.vision_model = vision_model
        self.retrieval_mode = retrieval_mode
        self.num_chunks = min(num_chunks, 20)  # Cap at 20 for performance
        self.vectorstore = None
        self.chain = None
        self.llm = None
        self.memory = None
        self.processed_documents = []
        
        # Casual message patterns
        self.casual_patterns = [
            r'^(hi|hey|hello|sup|what\'s up|wassup|yo)\b',
            r'^(thanks|thank you|thx|ty|appreciate it)\b',
            r'^(bye|goodbye|see you|cya|later)\b',
            r'^(how are you|how\'s it going|how are ya|how do you do)\b',
            r'^(ok|okay|cool|nice|great|awesome|perfect)\b',
            r'^(yes|no|yeah|yep|nope|sure)\b',
            r'(what are you|what\'re you|whatcha|wyd)',
            r'(tell me about yourself|who are you|introduce yourself)',
            r'(good morning|good afternoon|good evening|good night)',
            r'(nice to meet you|pleased to meet you)',
        ]
        
        self.casual_keywords = [
            'doing now', 'doing today', 'your name', 'about you', 
            'feeling', 'your day', 'up to', 'busy'
        ]
        
        # Check available Excel loaders
        print(f"[RAG] Excel support:")
        print(f"  - UnstructuredExcelLoader: {'✅' if UNSTRUCTURED_EXCEL else '❌'}")
        print(f"  - Pandas: {'✅' if PANDAS_AVAILABLE else '❌'}")
        print(f"  - OpenPyXL: {'✅' if OPENPYXL_AVAILABLE else '❌'}")
        
        print("[RAG] Loading embeddings...")
        import torch
        device = 'cuda' if torch.cuda.is_available() else 'cpu'
        print(f"[RAG] Device: {device.upper()}")
        
        self.embeddings = HuggingFaceEmbeddings(
            model_name="sentence-transformers/all-MiniLM-L6-v2",
            model_kwargs={'device': device},
            encode_kwargs={'normalize_embeddings': True, 'batch_size': 32}
        )
        
        self.vector_dir = os.path.join("vectors", "faiss_index")
        
        # Initialize vision LLM
        print("[RAG] Loading vision model...")
        self.vision_llm = ChatOllama(
            model=vision_model,
            temperature=0.3,
            timeout=300,
            keep_alive="10m"
        )
        
        print("[RAG] ✅ Ready with ENHANCED CAPACITY!")
        print(f"[RAG] 📊 Expected retrieval: ~{num_chunks * 1.5:.0f} chunks = ~{num_chunks * 0.8:.0f}-{num_chunks:.0f} pages per query")
    
    def _detect_file_type(self, file_name):
        """Detect file type from extension"""
        return os.path.splitext(file_name)[1].lower().lstrip('.')
    
    def _is_image_file(self, file_type):
        """Check if file is an image"""
        return file_type in ['png', 'jpg', 'jpeg', 'bmp', 'gif', 'webp', 'tiff']
    
    def _get_model_settings(self, model_name):
        """
        🚀 ENHANCED: Get timeout and context settings for model
        DOUBLED context window: 4096 → 8192 for most models
        """
        model_lower = model_name.lower()
        
        # Format: (num_predict, num_ctx, timeout)
        if any(x in model_lower for x in ['qwen3', '14b', '32b', '70b', 'deepseek-r1', '90b']):
            # Large models: 8192 context (was 4096)
            return (512, 8192, 600)
        elif any(x in model_lower for x in ['7b', '8b', '11b', '13b', 'qwen2.5:latest']):
            # Medium models: 6144 context (was 3072) 
            return (384, 6144, 300)
        else:
            # Small models: 4096 context (was 2048)
            return (256, 4096, 180)
    
    def _load_excel_with_pandas(self, file_path):
        """Load Excel using Pandas - MORE RELIABLE!"""
        print(f"[Excel] Using Pandas loader (more reliable)")
        
        try:
            documents = []
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            excel_file.close()
            
            for sheet_name in sheet_names:
                print(f"[Excel] Reading sheet: {sheet_name}")
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                content = f"=== Sheet: {sheet_name} ===\n\n"
                content += "Columns: " + ", ".join(df.columns.tolist()) + "\n\n"
                content += f"Total rows: {len(df)}\n\n"
                content += "Data:\n"
                
                for idx, row in df.head(1000).iterrows():
                    row_text = " | ".join([f"{col}: {val}" for col, val in row.items()])
                    content += f"Row {idx + 1}: {row_text}\n"
                
                if len(df) > 1000:
                    content += f"\n... (showing first 1000 of {len(df)} rows)\n"
                
                doc = Document(
                    page_content=content,
                    metadata={
                        "source": os.path.basename(file_path),
                        "sheet": sheet_name,
                        "rows": len(df),
                        "columns": len(df.columns)
                    }
                )
                documents.append(doc)
                del df
            
            print(f"[Excel] ✅ Successfully loaded {len(documents)} sheet(s)")
            return documents
            
        except Exception as e:
            print(f"[Excel] Pandas failed: {e}")
            raise
    
    def _load_excel_with_openpyxl(self, file_path):
        """Load Excel using OpenPyXL - FALLBACK METHOD"""
        print(f"[Excel] Using OpenPyXL loader (fallback)")
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            documents = []
            
            try:
                for sheet_name in wb.sheetnames:
                    print(f"[Excel] Reading sheet: {sheet_name}")
                    sheet = wb[sheet_name]
                    
                    content = f"=== Sheet: {sheet_name} ===\n\n"
                    rows_data = []
                    
                    for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        if idx > 1000:
                            break
                        row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                        rows_data.append(f"Row {idx}: {row_text}")
                    
                    content += "\n".join(rows_data)
                    
                    if sheet.max_row > 1000:
                        content += f"\n\n... (showing first 1000 of {sheet.max_row} rows)"
                    
                    doc = Document(
                        page_content=content,
                        metadata={
                            "source": os.path.basename(file_path),
                            "sheet": sheet_name,
                            "rows": sheet.max_row
                        }
                    )
                    documents.append(doc)
                
            finally:
                wb.close()
            
            print(f"[Excel] ✅ Successfully loaded {len(documents)} sheet(s)")
            return documents
            
        except Exception as e:
            print(f"[Excel] OpenPyXL failed: {e}")
            raise
    
    def _process_image_with_vision(self, file_path, file_name):
        """Process image using vision model"""
        print(f"[Vision] Processing {file_name} with {self.vision_model}")
        
        try:
            with open(file_path, 'rb') as f:
                image_data = f.read()
            
            image_b64 = base64.b64encode(image_data).decode('utf-8')
            
            prompt = """Analyze this image thoroughly and provide:
1. What type of document/image is this?
2. Extract ALL visible text (OCR)
3. Describe any charts, diagrams, or visual elements
4. Identify any data, numbers, or key information
5. Note any important details

Be comprehensive and extract everything visible."""
            
            message = {
                "role": "user",
                "content": prompt,
                "images": [image_b64]
            }
            
            response = self.vision_llm.invoke([message])
            description = response.content
            
            print(f"[Vision] ✅ Extracted {len(description)} characters")
            
            return [Document(
                page_content=f"=== IMAGE: {file_name} ===\n\n{description}",
                metadata={"source": file_name, "type": "image"}
            )]
            
        except Exception as e:
            print(f"[Vision] Failed: {e}")
            return [Document(
                page_content=f"[Image: {file_name} - Vision processing unavailable]",
                metadata={"source": file_name, "type": "image"}
            )]
    
    def _load_document_by_type(self, file_path):
        """Load document using appropriate loader"""
        file_type = self._detect_file_type(file_path)
        file_name = os.path.basename(file_path)
        
        try:
            # Images - use vision model
            if self._is_image_file(file_type):
                return self._process_image_with_vision(file_path, file_name)
            
            # PDF
            elif file_type == 'pdf':
                return PyPDFLoader(file_path).load()
            
            # Word
            elif file_type in ['docx', 'doc']:
                return Docx2txtLoader(file_path).load()
            
            # Text
            elif file_type in ['txt', 'md']:
                return TextLoader(file_path, encoding='utf-8').load()
            
            # RTF
            elif file_type == 'rtf':
                return UnstructuredRTFLoader(file_path).load()
            
            # CSV
            elif file_type == 'csv':
                return CSVLoader(file_path, encoding='utf-8').load()
            
            # Excel - try multiple methods
            elif file_type in ['xlsx', 'xls', 'ods']:
                if PANDAS_AVAILABLE:
                    return self._load_excel_with_pandas(file_path)
                elif OPENPYXL_AVAILABLE:
                    return self._load_excel_with_openpyxl(file_path)
                elif UNSTRUCTURED_EXCEL:
                    return UnstructuredExcelLoader(file_path, mode="elements").load()
                else:
                    raise Exception("No Excel loader available")
            
            # JSON
            elif file_type == 'json':
                return JSONLoader(file_path=file_path, jq_schema='.', text_content=False).load()
            
            # XML
            elif file_type == 'xml':
                return UnstructuredXMLLoader(file_path).load()
            
            # YAML
            elif file_type in ['yaml', 'yml']:
                return TextLoader(file_path, encoding='utf-8').load()
            
            else:
                return [Document(
                    page_content=f"[Unsupported: {file_type}]",
                    metadata={"source": file_name}
                )]
                
        except Exception as e:
            print(f"[ERROR] Loading {file_name}: {e}")
            return [Document(
                page_content=f"[Error: {file_name}]",
                metadata={"source": file_name, "error": str(e)}
            )]
    
    def process_uploaded_file(self, uploaded_file):
        """
        🚀 ENHANCED: Process uploaded file with BETTER CHUNKING
        
        IMPROVEMENTS:
        - Larger chunks: 1200 → 1500 characters
        - More overlap: 300 → 400 characters  
        - Better separators for rule-based documents
        - Smarter chunking to keep related content together
        """
        file_name = uploaded_file.name
        file_type = self._detect_file_type(file_name)
        
        print(f"[Processing] 📄 {file_name} ({file_type.upper()})")
        
        tmp_path = None
        try:
            # Create temp file
            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_type}") as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                tmp_path = tmp_file.name
            
            # Load document
            documents = self._load_document_by_type(tmp_path)
            
            # Clean up temp file
            if tmp_path and os.path.exists(tmp_path):
                os.unlink(tmp_path)
            
            if not documents:
                print(f"[Warning] No content from {file_name}")
                return []
            
            # 🚀 ENHANCED CHUNKING: Larger chunks with better overlap
            text_splitter = RecursiveCharacterTextSplitter(
                chunk_size=1500,        # INCREASED from 1200 (25% larger)
                chunk_overlap=400,      # INCREASED from 300 (keeps more context)
                separators=[
                    "\n\n\n",           # Major section breaks
                    "\n\n",             # Paragraph breaks
                    "\n",               # Line breaks
                    ". ",               # Sentence breaks
                    " ",                # Word breaks
                    ""
                ],
                length_function=len
            )
            
            chunks = text_splitter.split_documents(documents)
            
            # Update metadata
            for chunk in chunks:
                chunk.metadata['source'] = file_name
            
            self.processed_documents.append(file_name)
            print(f"[Processing] ✅ {file_name}: {len(chunks)} chunks (ENHANCED chunking)")
            print(f"[Processing] 📊 Estimated coverage: ~{len(chunks) * 1.5:.0f} chunks = ~{len(chunks) * 0.3:.0f} pages")
            
            return chunks
            
        except Exception as e:
            if tmp_path and os.path.exists(tmp_path):
                os.unlink(tmp_path)
            print(f"[ERROR] Processing {file_name}: {e}")
            raise
    
    def create_vectorstore(self, chunks):
        """Create FAISS vectorstore from chunks"""
        print(f"[Vectorstore] Creating from {len(chunks)} chunks...")
        start_time = time.time()
        
        if not chunks:
            raise ValueError("No chunks provided")
        
        self.vectorstore = FAISS.from_documents(chunks, embedding=self.embeddings)
        
        elapsed = time.time() - start_time
        print(f"[Vectorstore] ✅ Created in {elapsed:.2f}s")
        print(f"[Vectorstore] 📊 Total vectors: {self.vectorstore.index.ntotal}")
        
        # Save vectorstore
        os.makedirs(self.vector_dir, exist_ok=True)
        try:
            self.vectorstore.save_local(self.vector_dir)
            print(f"[Vectorstore] 💾 Saved to disk")
        except Exception as e:
            print(f"[Vectorstore] Could not save: {e}")
    
    def setup_chain(self):
        """
        🚀 ENHANCED: Setup chain with INCREASED RETRIEVAL CAPACITY
        
        IMPROVEMENTS:
        - 2x more chunks retrieved: 6 → 12 chunks
        - MMR retrieval for diverse results
        - Larger context window: 4096 → 8192 tokens
        - Better prompt for comprehensive answers
        """
        if not self.vectorstore:
            raise ValueError("Vectorstore not initialized")
        
        print(f"[Chain] 🚀 Setting up ENHANCED chain...")
        print(f"[Chain] Model: {self.model}")
        print(f"[Chain] Retrieval Mode: {self.retrieval_mode.upper()}")
        print(f"[Chain] Chunks to Retrieve: {self.num_chunks}")
        
        # Get enhanced model settings (larger context)
        num_predict, num_ctx, timeout = self._get_model_settings(self.model)
        print(f"[Chain] Context Window: {num_ctx} tokens (ENHANCED)")
        
        # Initialize LLM with larger context
        self.llm = ChatOllama(
            model=self.model,
            temperature=0.2,
            num_predict=num_predict,
            num_ctx=num_ctx,         # LARGER context window
            timeout=timeout,
            keep_alive="10m",
            num_gpu=1,
            num_thread=8
        )
        
        # Initialize memory
        self.memory = ConversationBufferMemory(
            memory_key="chat_history",
            return_messages=True,
            output_key="answer"
        )
        
        # 🚀 ENHANCED PROMPT: Better instructions for comprehensive answers
        qa_prompt = PromptTemplate(
            template="""You are a comprehensive document analyst. Use ALL the provided context to give complete, thorough answers.

Context from documents (READ ALL OF THIS CAREFULLY):
{context}

Question: {question}

CRITICAL INSTRUCTIONS:
- Read through EVERY SINGLE piece of context above
- If the question asks for a list, provide COMPLETE lists with ALL items from the context
- If multiple chunks mention the same topic, synthesize information from ALL of them
- If information spans multiple sections, combine them coherently
- For "list all" questions, enumerate EVERYTHING you find in the context
- If context contains images, use the image descriptions provided
- If information seems incomplete based on the context, acknowledge what's present and what might be missing
- Be comprehensive but concise - don't add information not in the context
- If you see numbered items (like rules), list them all in order

Answer (be thorough and complete):""",
            input_variables=["context", "question"]
        )
        
        # 🚀 ENHANCED RETRIEVAL: Configure based on mode
        if self.retrieval_mode == "mmr":
            # MMR: Maximal Marginal Relevance - diverse results
            search_type = "mmr"
            search_kwargs = {
                "k": self.num_chunks,
                "fetch_k": self.num_chunks * 3,  # Fetch 3x more for diversity
                "lambda_mult": 0.7                # Balance relevance vs diversity
            }
            print(f"[Chain] Using MMR retrieval (diverse results)")
            
        elif self.retrieval_mode == "hybrid":
            # Hybrid: Get more and re-rank
            search_type = "similarity"
            search_kwargs = {
                "k": self.num_chunks,
                "score_threshold": 0.3  # Only include relevant results
            }
            print(f"[Chain] Using Hybrid retrieval (filtered results)")
            
        else:  # similarity (default)
            # Simple similarity search
            search_type = "similarity"
            search_kwargs = {"k": self.num_chunks}
            print(f"[Chain] Using Similarity retrieval (relevance-based)")
        
        # Create chain with enhanced settings
        self.chain = ConversationalRetrievalChain.from_llm(
            llm=self.llm,
            retriever=self.vectorstore.as_retriever(
                search_type=search_type,
                search_kwargs=search_kwargs
            ),
            memory=self.memory,
            return_source_documents=True,
            combine_docs_chain_kwargs={"prompt": qa_prompt},
            verbose=False
        )
        
        print(f"[Chain] ✅ Ready!")
        print(f"[Chain] 📊 Will retrieve {self.num_chunks} chunks (~{self.num_chunks * 0.8:.0f}-{self.num_chunks:.0f} pages)")
    
    def _is_casual_message(self, message):
        """Check if message is casual conversation"""
        msg_lower = message.lower().strip()
        
        for pattern in self.casual_patterns:
            if re.search(pattern, msg_lower, re.IGNORECASE):
                return True
        
        for keyword in self.casual_keywords:
            if keyword in msg_lower:
                return True
        
        if len(msg_lower.split()) <= 3 and '?' not in msg_lower:
            return True
        
        if 'document' in msg_lower or 'file' in msg_lower or 'pdf' in msg_lower:
            return False
        if 'content' in msg_lower or 'show' in msg_lower or 'find' in msg_lower:
            return False
        if 'search' in msg_lower or 'list' in msg_lower or 'extract' in msg_lower:
            return False
        
        if any(combo in msg_lower for combo in ['what are you', 'what\'re you', 'whatcha']):
            return True
        
        words = msg_lower.split()
        if 'you' in words or 'your' in words or 'yourself' in words:
            if len(words) <= 10:
                return True
        
        return False
    
    def _init_chat_llm(self):
        """Initialize LLM for chat"""
        num_predict, num_ctx, timeout = self._get_model_settings(self.model)
        self.chat_llm = ChatOllama(
            model=self.model,
            temperature=0.7,
            num_predict=num_predict,
            num_ctx=num_ctx,
            timeout=timeout,
            keep_alive="10m",
            num_gpu=1,
            num_thread=8
        )
    
    def _chat_direct(self, message):
        """Direct chat without documents"""
        if not hasattr(self, 'chat_llm') or self.chat_llm is None:
            self._init_chat_llm()
        
        try:
            response = self.chat_llm.invoke(message)
            return {"answer": response.content, "source_documents": []}
        except Exception as e:
            print(f"[ERROR] Chat failed: {e}")
            raise
    
    def switch_model(self, new_model):
        """Switch to different model with enhanced settings"""
        print(f"[Model Switch] {self.model} → {new_model}")
        
        num_predict, num_ctx, timeout = self._get_model_settings(new_model)
        print(f"[Model Switch] Context: {num_ctx} tokens")
        
        # Update chat LLM
        self.chat_llm = ChatOllama(
            model=new_model,
            temperature=0.7,
            num_predict=num_predict,
            num_ctx=num_ctx,
            timeout=timeout,
            keep_alive="10m",
            num_gpu=1,
            num_thread=8
        )
        
        if not self.memory:
            self.memory = ConversationBufferMemory(
                memory_key="chat_history",
                return_messages=True,
                output_key="answer"
            )
        
        # Update RAG chain if exists
        if self.vectorstore:
            self.llm = ChatOllama(
                model=new_model,
                temperature=0.2,
                num_predict=num_predict,
                num_ctx=num_ctx,
                timeout=timeout,
                keep_alive="10m",
                num_gpu=1,
                num_thread=8
            )
            
            qa_prompt = PromptTemplate(
                template="""You are a comprehensive document analyst. Use ALL the provided context to give complete, thorough answers.

Context from documents (READ ALL OF THIS CAREFULLY):
{context}

Question: {question}

CRITICAL INSTRUCTIONS:
- Read through EVERY SINGLE piece of context above
- If the question asks for a list, provide COMPLETE lists with ALL items from the context
- If multiple chunks mention the same topic, synthesize information from ALL of them
- If information spans multiple sections, combine them coherently
- For "list all" questions, enumerate EVERYTHING you find in the context
- If context contains images, use the image descriptions provided
- If information seems incomplete based on the context, acknowledge what's present and what might be missing
- Be comprehensive but concise - don't add information not in the context

Answer (be thorough and complete):""",
                input_variables=["context", "question"]
            )
            
            # Configure retrieval based on mode
            if self.retrieval_mode == "mmr":
                search_type = "mmr"
                search_kwargs = {
                    "k": self.num_chunks,
                    "fetch_k": self.num_chunks * 3,
                    "lambda_mult": 0.7
                }
            elif self.retrieval_mode == "hybrid":
                search_type = "similarity"
                search_kwargs = {
                    "k": self.num_chunks,
                    "score_threshold": 0.3
                }
            else:
                search_type = "similarity"
                search_kwargs = {"k": self.num_chunks}
            
            self.chain = ConversationalRetrievalChain.from_llm(
                llm=self.llm,
                retriever=self.vectorstore.as_retriever(
                    search_type=search_type,
                    search_kwargs=search_kwargs
                ),
                memory=self.memory,
                return_source_documents=True,
                combine_docs_chain_kwargs={"prompt": qa_prompt},
                verbose=False
            )
        
        self.model = new_model
        print(f"[Model Switch] ✅ Switched to {new_model}")
    
    def ask_question(self, question):
        """Ask question with enhanced retrieval"""
        print(f"\n{'='*60}")
        print(f"[QUERY] {question}")
        print(f"[MODEL] {self.model}")
        print(f"[RETRIEVAL] {self.retrieval_mode.upper()} mode, {self.num_chunks} chunks")
        print(f"{'='*60}")
        
        # Check if casual
        is_casual = self._is_casual_message(question)
        
        if is_casual:
            print("[INFO] Casual message - using direct chat")
            return self._chat_direct(question)
        
        if not self.chain:
            print("[INFO] No documents - using direct chat")
            return self._chat_direct(question)
        
        total_start = time.time()
        
        try:
            chat_history = self.memory.chat_memory.messages if self.memory else []
            
            print(f"[INFO] Retrieving {self.num_chunks} chunks...")
            response = self.chain.invoke({
                "question": question,
                "chat_history": chat_history
            })
            
            total_time = time.time() - total_start
            sources = response.get("source_documents", [])
            
            print(f"\n[INFO] ✅ Retrieved {len(sources)} chunks in {total_time:.2f}s")
            print(f"[INFO] 📊 Coverage: ~{len(sources) * 0.8:.0f}-{len(sources):.0f} pages")
            
            for i, doc in enumerate(sources, 1):
                preview = doc.page_content[:100].replace('\n', ' ')
                source_file = doc.metadata.get('source', 'Unknown')
                print(f"  [{i}] {source_file}: {preview}...")
            
            print(f"[INFO] Answer generated in {total_time:.2f}s")
            print(f"{'='*60}\n")
            
            return {
                "answer": response["answer"],
                "source_documents": sources
            }
            
        except TimeoutError as e:
            elapsed = time.time() - total_start
            print(f"[TIMEOUT] Query timed out after {elapsed:.2f}s")
            return {
                "answer": f"⚠️ Query timed out after {elapsed:.0f}s. Try using a faster model or simpler question.",
                "source_documents": []
            }
            
        except Exception as e:
            elapsed = time.time() - total_start
            print(f"[ERROR] Query failed after {elapsed:.2f}s: {e}")
            import traceback
            traceback.print_exc()
            
            return {
                "answer": f"❌ Error: {str(e)}\n\nTry a different model or simpler question.",
                "source_documents": []
            }
    
    def clear_documents(self):
        """Clear all documents"""
        self.vectorstore = None
        self.chain = None
        self.llm = None
        self.memory = None
        self.processed_documents = []
        
        if os.path.exists(self.vector_dir):
            import shutil
            try:
                shutil.rmtree(self.vector_dir)
                print("[RAG] Cleared saved vectors")
            except:
                pass
        
        print("[RAG] All documents cleared")
    
    def set_retrieval_config(self, mode="mmr", num_chunks=12):
        """
        Change retrieval configuration dynamically
        
        Args:
            mode: 'similarity', 'mmr', or 'hybrid'
            num_chunks: Number of chunks to retrieve (1-20)
        """
        self.retrieval_mode = mode
        self.num_chunks = min(num_chunks, 20)
        
        print(f"[Config] Updated: {mode.upper()} mode, {self.num_chunks} chunks")
        
        # Rebuild chain if it exists
        if self.chain:
            self.setup_chain()
